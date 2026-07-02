"""
BA Intelligence Toolkit — Utilities
File I/O, Excel export, and visualization helpers.
"""

import io
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------

def load_text_file(file_path: str | Path) -> str:
    """Load a text file and return its contents."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def load_uploaded_file(uploaded_file) -> str:
    """Load text from a Streamlit UploadedFile object."""
    if uploaded_file is None:
        return ""
    # Handle different file types
    name = uploaded_file.name.lower()
    if name.endswith(".txt"):
        return uploaded_file.read().decode("utf-8")
    elif name.endswith(".pdf"):
        # Basic PDF text extraction
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(uploaded_file.read()))
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    else:
        # Try reading as text
        try:
            return uploaded_file.read().decode("utf-8")
        except UnicodeDecodeError:
            return uploaded_file.read().decode("latin-1")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(
    requirements: list[dict] | None = None,
    compliance_report: dict | None = None,
    rtm_data: dict | None = None,
    gap_analysis: dict | None = None,
) -> bytes:
    """Export all results to an Excel file (in-memory bytes).

    Each section becomes a worksheet.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Requirements sheet ---
    if requirements:
        ws = wb.create_sheet("Requirements")
        df = pd.DataFrame(requirements)
        _write_dataframe(ws, df)

    # --- Compliance sheet ---
    if compliance_report and compliance_report.get("raw_results"):
        ws = wb.create_sheet("Compliance Gap Check")
        rows = []
        for r in compliance_report["raw_results"]:
            rows.append({
                "Obligation ID": r.get("obligation_id", ""),
                "Status": r.get("status", ""),
                "Reasoning": r.get("reasoning", ""),
                "Consequence if Gap": r.get("consequence_if_gap", ""),
                "Suggested Control": r.get("suggested_control", ""),
            })
        df = pd.DataFrame(rows)
        _write_dataframe(ws, df)

    # --- RTM sheet ---
    if rtm_data and rtm_data.get("rtm_entries"):
        ws = wb.create_sheet("RTM")
        df = pd.DataFrame(rtm_data["rtm_entries"])
        _write_dataframe(ws, df)

    # --- Gap Analysis sheet ---
    if gap_analysis and gap_analysis.get("gaps"):
        ws = wb.create_sheet("Gap Analysis")
        df = pd.DataFrame(gap_analysis["gaps"])
        _write_dataframe(ws, df)

    # --- Summary sheet (always) ---
    ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
    summary_lines = ["BA Intelligence Toolkit — Export Summary", ""]
    if requirements:
        summary_lines.append(f"Requirements extracted: {len(requirements)}")
    if compliance_report:
        s = compliance_report.get("summary", {})
        summary_lines.append(
            f"Compliance: {s.get('satisfied', 0)} satisfied, "
            f"{s.get('gaps', 0)} gaps, {s.get('unclear', 0)} unclear"
        )
        if s.get("high_risk_gaps"):
            summary_lines.append(f"  High-risk gaps: {s['high_risk_gaps']}")
    if gap_analysis:
        summary_lines.append(
            f"Process gaps identified: {len(gap_analysis.get('gaps', []))}"
        )
    for i, line in enumerate(summary_lines, 1):
        ws.cell(row=i, column=1, value=line)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_dataframe(ws, df: pd.DataFrame):
    """Write a DataFrame to a worksheet with headers."""
    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = cell.font.copy(bold=True)
    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")


# ---------------------------------------------------------------------------
# Visualization helpers
# ---------------------------------------------------------------------------

def create_compliance_heatmap(report: dict) -> go.Figure:
    """Create a heatmap visualization of compliance check results."""
    results = report.get("raw_results", [])
    if not results:
        fig = go.Figure()
        fig.update_layout(title="No compliance data to display")
        return fig

    # Group by status
    statuses = ["satisfied", "gap", "unclear", "not_applicable"]
    colors = {"satisfied": "#2ecc71", "gap": "#e74c3c",
              "unclear": "#f39c12", "not_applicable": "#95a5a6"}

    ids = [r.get("obligation_id", "?") for r in results]
    status_list = [r.get("status", "?") for r in results]
    color_list = [colors.get(s, "#999") for s in status_list]

    fig = go.Figure(data=go.Bar(
        x=ids,
        y=[1] * len(ids),
        marker_color=color_list,
        text=status_list,
        textposition="auto",
        hovertemplate="<b>%{x}</b><br>Status: %{text}<extra></extra>",
    ))
    fig.update_layout(
        title="Compliance Obligation Check Results",
        xaxis_title="Obligation ID",
        yaxis=dict(visible=False),
        height=400,
        showlegend=False,
    )
    return fig


def create_gap_priority_matrix(gap_analysis: dict) -> go.Figure:
    """Create a priority matrix scatter plot for process gaps."""
    matrix = gap_analysis.get("priority_matrix", [])
    if not matrix:
        fig = go.Figure()
        fig.update_layout(title="No gap priority data to display")
        return fig

    impact_map = {"High": 3, "Medium": 2, "Low": 1}
    difficulty_map = {"High": 3, "Medium": 2, "Low": 1}

    x_vals = [difficulty_map.get(m.get("difficulty", "Medium"), 2) for m in matrix]
    y_vals = [impact_map.get(m.get("impact", "Medium"), 2) for m in matrix]
    labels = [m.get("gap_id", "?") for m in matrix]

    fig = go.Figure(data=go.Scatter(
        x=x_vals,
        y=y_vals,
        mode="markers+text",
        text=labels,
        textposition="top center",
        marker=dict(size=15, color="#3498db"),
        hovertemplate="<b>%{text}</b><br>Impact: %{y}<br>Difficulty: %{x}<extra></extra>",
    ))
    fig.update_layout(
        title="Gap Priority Matrix (Impact vs Difficulty)",
        xaxis=dict(
            title="Implementation Difficulty",
            tickvals=[1, 2, 3],
            ticktext=["Low", "Medium", "High"],
        ),
        yaxis=dict(
            title="Impact",
            tickvals=[1, 2, 3],
            ticktext=["Low", "Medium", "High"],
        ),
        height=450,
    )
    return fig
